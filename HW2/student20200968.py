#!/usr/bin/python3

import openpyxl

wb = openpyxl.load_workbook("student.xlsx")
ws = wb['Sheet1']

row_id = 1;
for row in ws:
    if row_id != 1:
        sum_v = ws.cell(row = row_id, column = 3).value * 0.3
        sum_v += ws.cell(row = row_id, column = 4).value * 0.35
        sum_v += ws.cell(row = row_id, column = 5).value * 0.34
        sum_v += ws.cell(row = row_id, column = 6).value 
        ws.cell(row = row_id, column = 7).value = sum_v
    row_id += 1

total = row_id - 2
A = int(total * 0.3)
Ap = int(A * 0.5)
B = int(total * 0.7)
Bp = int((B - A) * 0.5)
C = total - B
Cp = int(C * 0.5)
rank = [] 
grade = []
ag = []
row_id = 2
i = 0; j = 0

for row in ws:
    ag.append(ws.cell(row = row_id, column = 7).value)
    row_id += 1

for i in range(len(ag) - 1 ):
    rank.insert(i,1)
    for j in range(len(ag) - 1):
        if ag[i] < ag[j]:
            rank[i] += 1
        j += 1
    i += 1

i = 0
for i in range(len(rank)):
    if A >= rank[i]:
        if Ap >= rank[i]:
            rank[i] = 'A+'
        else:
            rank[i] = 'A0'
    elif B >= rank[i]:
        if Bp + A  >= rank[i]:
            rank[i] = 'B+'
        else:
            rank[i] = 'B0'
    else:
        if Cp + B >= rank[i]:
            rank[i] = 'C+'
        else:
            rank[i] = 'C0'

i = 0
for row in range(len(rank)): 
    ws.cell(row = i + 2, column = 8).value = rank[i]
    i += 1

wb.save("student.xlsx")

